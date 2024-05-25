import os
import time

import openpyxl
import pandas as pd
from memory_profiler import memory_usage
from openpyxl.chart import LineChart, Reference

from Sudoku import get_random_sudoku
from exact_algorithm import ExactAlgorithm
from genetic_algorithm import GeneticAlgorithm
from source_manager import get_solutions


class PerformanceRunner:
    """ PerformanceRunner class to run preconfigured performance calculations and save them with analysis to excel """

    def __init__(self, num_of_solutions_to_use, attempts_amount, num_of_empty_cells_range,
                 genetic_algorithm_population_numbers, output_file_name):
        self.attempts_amount = attempts_amount
        self.num_of_empty_cells_range = num_of_empty_cells_range
        self.genetic_algorithm_population_numbers = genetic_algorithm_population_numbers

        if not os.path.exists('performance_results'):
            os.makedirs('performance_results')
        self.output_file_path = os.path.join('performance_results', output_file_name)

        self.good_solutions = get_solutions(num_of_solutions_to_use)

        self.average_summary = {
            solver: [] for solver in
            ['exact_time'] +
            ['exact_memory'] +
            [f'genetic_{p}_time' for p in self.genetic_algorithm_population_numbers] +
            [f'genetic_{p}_memory' for p in self.genetic_algorithm_population_numbers]
        }

        self.results = {}
        for n in self.num_of_empty_cells_range:
            self.results.update({f'index_{n}': []})
            self.results.update({f'exact_{n}_time': []})
            self.results.update({f'exact_{n}_memory': []})
            self.results.update({f'genetic_{p}_{n}_time': [] for p in self.genetic_algorithm_population_numbers})
            self.results.update({f'genetic_{p}_{n}_memory': [] for p in self.genetic_algorithm_population_numbers})

    @staticmethod
    def run_exact_algorithm_performance(good_solution, num_of_empty_cells):
        """ exact algorithm runner with time and memory measure """

        print(f"Running ExactAlgorithm for n={num_of_empty_cells}")

        def run_algorithm():
            """ run  exact_algorithm """
            sudoku = get_random_sudoku(num_of_empty_cells, good_solution)
            exact_algorithm = ExactAlgorithm(sudoku)

            start_time = time.time()

            result = exact_algorithm.exact_algorithm()

            end_time = time.time()

            return result, end_time - start_time

        memory_usage_info, run_result = memory_usage(run_algorithm, interval=0.1, max_usage=True, retval=True)

        result_sudoku, elapsed_time = run_result

        if not result_sudoku.is_valid():
            raise ValueError("ExactAlgorithm: Sudoku not solved properly")

        print(f"Finished ExactAlgorithm for n={num_of_empty_cells} "
              f"with time {elapsed_time} seconds and max memory usage {memory_usage_info} MiB")
        return elapsed_time, memory_usage_info

    @staticmethod
    def run_genetic_algorithm_performance(good_solution, num_of_empty_cells, population_number):
        """ genetic algorithm runner with time and memory measure"""

        print(f"Running GeneticAlgorithm for p={population_number} and n={num_of_empty_cells}")

        def run_algorithm():
            """ run  genetic_algorithm """
            sudoku = get_random_sudoku(num_of_empty_cells, good_solution)
            genetic_algorithm = GeneticAlgorithm(sudoku, population_number=population_number)

            start_time = time.time()

            result = genetic_algorithm.genetic_algorithm()

            end_time = time.time()

            return result, end_time - start_time

        memory_usage_info, run_result = memory_usage(run_algorithm, interval=0.1, max_usage=True, retval=True)

        result_sudoku, elapsed_time = run_result

        if not result_sudoku.is_valid():
            raise ValueError(f"GeneticAlgorithm {population_number}: Sudoku not solved properly")

        print(f"Finished GeneticAlgorithm for p={population_number} and n={num_of_empty_cells} "
              f"with time {elapsed_time} seconds and max memory usage {memory_usage_info} MiB")
        return elapsed_time, memory_usage_info

    def run_for_solution(self, good_solution):
        """ exact and genetic algorithms runner for different num_of_empty_cells """

        for num_of_empty_cells in self.num_of_empty_cells_range:
            print(f"Running for n={num_of_empty_cells}")

            for _ in range(self.attempts_amount):
                exact_time, exact_memory = self.run_exact_algorithm_performance(good_solution, num_of_empty_cells)
                self.results[f'exact_{num_of_empty_cells}_time'].append(exact_time)
                self.results[f'exact_{num_of_empty_cells}_memory'].append(exact_memory)

            for population_number in self.genetic_algorithm_population_numbers:
                for _ in range(self.attempts_amount):
                    genetic_time, genetic_memory = self.run_genetic_algorithm_performance(
                        good_solution, num_of_empty_cells, population_number)
                    self.results[f'genetic_{population_number}_{num_of_empty_cells}_time'].append(genetic_time)
                    self.results[f'genetic_{population_number}_{num_of_empty_cells}_memory'].append(genetic_memory)

            print(f"Finished for n={num_of_empty_cells}")

    def save_time_memory_results(self):
        """ save time results from running solvers with different solutions or number of empty cells """

        df = pd.DataFrame.from_dict(self.results, orient='index').transpose()
        for n in self.num_of_empty_cells_range:
            df[f'index_{n}'] = df.index

        averages_df = pd.DataFrame([df.mean()], index=['Average'])
        df = pd.concat([df, averages_df])

        with pd.ExcelWriter(self.output_file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name="Performance Results")

    def create_time_memory_charts(self):
        """ create charts for every number of empty cells to compare solvers time or memory performance """

        wb = openpyxl.load_workbook(self.output_file_path)
        ws = wb.active

        column_titles = {cell.value: cell.column for cell in ws[1]}

        time_config = {'title_part': 'Time', 'col_postfix': 'time', 'chart_col_placing': 'A'}
        memory_config = {'title_part': 'Memory Usage', 'col_postfix': 'memory', 'chart_col_placing': 'K'}
        for config in [time_config, memory_config]:

            chart_placement = 2

            for n in self.num_of_empty_cells_range:
                chart = LineChart()
                chart.title = f"{config['title_part']} Performance Comparison - {n} Empty Cells"
                chart.style = 13

                exact_column = column_titles[f'exact_{n}_{config['col_postfix']}']
                exact_data = Reference(ws, min_col=exact_column, min_row=1, max_row=ws.max_row - 1)
                chart.add_data(exact_data, titles_from_data=True)

                for p in self.genetic_algorithm_population_numbers:
                    genetic_column = column_titles[f'genetic_{p}_{n}_{config['col_postfix']}']
                    genetic_data = Reference(ws, min_col=genetic_column, min_row=1, max_row=ws.max_row - 1)
                    chart.add_data(genetic_data, titles_from_data=True)

                chart.legend.position = 'r'
                chart.x_axis.title = "Trials"
                chart.y_axis.title = f"{config['title_part']} Performance"

                ws.add_chart(chart, f"{config['chart_col_placing']}{ws.max_row + chart_placement}")
                chart_placement += 15

        wb.save(self.output_file_path)

    def get_avg_results(self):
        """ get average time results and save it in different sheet """

        wb = openpyxl.load_workbook(self.output_file_path)
        ws = wb.active

        column_titles = {cell.value: cell.column for cell in ws[1]}

        for solver in self.average_summary:
            for n in self.num_of_empty_cells_range:
                solver_col = solver
                if '_time' in solver:
                    solver_col = solver_col.replace('_time', f'_{n}_time')
                elif '_memory' in solver:
                    solver_col = solver_col.replace('_memory', f'_{n}_memory')
                if solver_col in column_titles:
                    col = column_titles[solver_col]
                    self.average_summary[solver].append(ws.cell(row=ws.max_row, column=col).value)

        wb.save(self.output_file_path)

        df_avg = pd.DataFrame.from_dict(self.average_summary, orient='index').transpose()

        df_avg.index = self.num_of_empty_cells_range

        with pd.ExcelWriter(self.output_file_path, engine='openpyxl', mode='a') as writer:
            df_avg.to_excel(writer, sheet_name='Average Performance Comparison')

    def create_avg_time_memory_chart(self):
        """ create charts for average times or memory usages of solvers to compare their performance """

        avg_wb = openpyxl.load_workbook(self.output_file_path)
        avg_ws = avg_wb['Average Performance Comparison']

        column_titles = {cell.value: cell.column for cell in avg_ws[1]}

        time_config = {'title_part': 'Time', 'col_postfix': 'time', 'chart_col_placing': 'A'}
        memory_config = {'title_part': 'Memory Usage', 'col_postfix': 'memory', 'chart_col_placing': 'K'}
        for config in [time_config, memory_config]:

            avg_chart = LineChart()
            avg_chart.title = f"Average {config['title_part']} Performance Comparison"
            avg_chart.style = 13

            for solver in [solver for solver in self.average_summary if f'_{config['col_postfix']}' in solver]:
                exact_column = column_titles[solver]
                exact_data = Reference(avg_ws, min_col=exact_column, min_row=1, max_row=avg_ws.max_row)
                avg_chart.add_data(exact_data, titles_from_data=True)

            categories = Reference(avg_ws, min_col=1, min_row=2, max_row=avg_ws.max_row)
            avg_chart.set_categories(categories)

            avg_chart.legend.position = 'r'
            avg_chart.x_axis.title = "Number of empty cells"
            avg_chart.y_axis.title = f"Average {config['title_part']} Performance"

            avg_ws.add_chart(avg_chart, f"{config['chart_col_placing']}{avg_ws.max_row + 2}")

        avg_wb.save(os.path.join('performance_results', 'performance_results.xlsx'))

    def main(self):
        """ main performance runner for different good solutions """

        for i, good_solution in enumerate(self.good_solutions):
            print(f"Running for good_solution {i + 1}")
            self.run_for_solution(good_solution)
            print(f"Finished for good_solution {i + 1}")

        self.save_time_memory_results()
        self.create_time_memory_charts()
        self.get_avg_results()
        self.create_avg_time_memory_chart()


if __name__ == '__main__':
    NUM_OF_SOLUTIONS_TO_USE = 3
    ATTEMPTS_AMOUNT = 5
    GENETIC_ALGORYTHM_POPULATION_NUMBERS = [5, 10, 20]
    NUM_OF_EMPTY_CELLS_RANGE = range(3, 7)
    OUTPUT_FILE_NAME = 'performance_results.xlsx'

    runner = PerformanceRunner(NUM_OF_SOLUTIONS_TO_USE, ATTEMPTS_AMOUNT, NUM_OF_EMPTY_CELLS_RANGE,
                               GENETIC_ALGORYTHM_POPULATION_NUMBERS, OUTPUT_FILE_NAME)
    runner.main()
